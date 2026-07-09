# ==============================================================================
# 1. IMPORTACIÓN DE LIBRERÍAS
# ==============================================================================
import logging                           # Para guardar registro de eventos (logs)
import smtplib                           # Para enviar correos electrónicos
import time                              # Para manejar tiempos y pausas
import os                                # Para limpiar pantalla del sistema
import psutil                            # Para leer el tráfico de la tarjeta de red
import csv                               # Para leer y escribir archivos en formato csv
import datetime                          # Para manejar y realizar operaciones con el tiempo
import socket                            # Para acceder a la tarjeta de red y obtener la ip local
import pandas as pd                      # Para analisis y manipulacion de datos
import mimetypes                         # Para identificar el tipo de imagen del logo
import matplotlib
matplotlib.use("Agg")                    # Backend sin interfaz gráfica (evita errores de Tkinter)
import matplotlib.pyplot as plt          # Para creacion de graficos
import platform
from email.message import EmailMessage   # Para dar formato al correo
from scapy.all import sniff, ARP, Ether, srp, conf   # Scapy: El núcleo del análisis de red
from PIL import Image                    # Pillow: Para procesar la imagen del logo

# ==============================================================================
# 2. CONFIGURACIONES Y COLORES
# ==============================================================================
ROJO = '\033[91m'             
VERDE = '\033[92m'            
AMARILLO = '\033[93m'         
AZUL = '\033[94m'             
RESET = '\033[0m'             
NEGRITA = '\033[1m'
ANCHO_CONSOLA = 60            

# --- CONFIGURACIÓN DE CORREO ---
SMTP_SERVER = "smtp.gmail.com"               
SMTP_PORT = 587                              
EMAIL_USER = "coblitasv2@gmail.com"          # <--- TU CORREO
EMAIL_PASS = "nyaj rouc jenj usuj"           # <--- TU CONTRASEÑA DE APP
EMAIL_DESTINO = "pcislcas@upc.edu.pe"      # <--- CORREO DESTINO
NOMBRE_LOGO = "descarga.png"                 # <--- ASEGURATE QUE ESTE ARCHIVO EXISTA

# --- CONFIGURACIÓN DE RED ---
IP_MAC_CONFIANZA = {
    "192.168.1.1" : "74:93:da:e8:44:80",  
    "192.168.1.42": "a8:3b:76:19:07:cd",  
    "192.168.1.45": "74:4c:a1:ca:1d:67",  
}
NOMBRE_INTERFAZ = "Wi-Fi" 

# --- CONFIGURACIÓN ESCÁNER ---
RANGO_RED_SCANNER = "192.168.1.0/24"

# --- CONFIGURACIÓN DE ALERTAS ---  
TIEMPO_ENTRE_ALERTAS = 10     
TIEMPO_ENTRE_ALERTAS_BW = 60  
UMBRAL_MBPS = 100             

# --- CONFIGURACIÓN PING ---
PING_ARCHIVO_SERVIDORES = "servidores.csv"
PING_ARCHIVO_HISTORIAL = "historial_ping.csv"
PING_ARCHIVO_RESUMEN = "ping_resumen.csv"
PING_SERVIDORES_DEFECTO = [
    ("Servidor AD",   "192.168.1.42"), 
    ("Servidor WEB",  "192.168.1.45"), 
    ("Puerta de Enlace",  "192.168.1.1"), 
    ("Servidor DNS",  "8.8.8.8")
]
PING_UMBRAL_EXCELENTE = 99.0
PING_UMBRAL_BUENO      = 95.0
PING_UMBRAL_ACEPTABLE  = 80.0

# Variables de estado
ultima_alerta_arp = 0         
ultima_alerta_bw = 0          

# Configuración de Logs
logging.basicConfig(filename='seguridad_sistema.log', level=logging.WARNING, 
                    format='%(asctime)s - %(message)s')

# Configuración Scapy
conf.verb = 0       
conf.resolve = []   

# Archivos de Ancho de Banda
ARCHIVO_CSV = 'historial_trafico.csv'
ARCHIVO_IMG = 'grafico_trafico.png'

# ==============================================================================
# 3. FUNCIONES AUXILIARES (UI y NET)
# ==============================================================================
def limpiar_pantalla():
    if os.name == 'nt': os.system('cls')
    else: os.system('clear') 

def convertir_bytes(peso):
    for unidad in ['B', 'KB', 'MB', 'GB']:
        if peso < 1024.0:                
            return f"{peso:.2f} {unidad}"
        peso /= 1024.0

def obtener_ip_local():
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 1))
        IP = s.getsockname()[0]
    except Exception:
        IP = '127.0.0.1' 
    finally:
        s.close()
    return IP

def crear_caratula(ruta_imagen, ancho_logo=30):
    limpiar_pantalla()
    logo_ascii = []
    try:
        img = Image.open(ruta_imagen)
        ancho_orig, alto_orig = img.size
        relacion = alto_orig / ancho_orig
        alto_logo = int(ancho_logo * relacion * 0.55)
        img = img.resize((ancho_logo, alto_logo)).convert("L")
        pixeles = img.getdata()
        cadena_ascii = "".join([" " if p > 150 else "█" for p in pixeles])
        logo_ascii = [cadena_ascii[i:i+ancho_logo] for i in range(0, len(cadena_ascii), ancho_logo)]
    except Exception: pass

    print("\n" + "=" * ANCHO_CONSOLA)
    print(f"{NEGRITA}{'TRABAJO FINAL: FUNDAMENTOS DE PROGRAMACIÓN'.center(ANCHO_CONSOLA)}{RESET}")
    print("-" * ANCHO_CONSOLA)
    print("Profesor: Luis Castillo Mesias".center(ANCHO_CONSOLA))
    print("NRC:803".center(ANCHO_CONSOLA))
    print("Carrera de Ingeniería de Redes y Comunicaciones".center(ANCHO_CONSOLA))
    print(" " * ANCHO_CONSOLA)
    print("Grupo 3".center(ANCHO_CONSOLA))
    print("Integrantes: \n -Carlos Jesus Oblitas Vargas\n -Madai Gonzales Collantes\n -Raphael Enrique Avalos Espinoza\n -Bils Ortega Buitron".center(ANCHO_CONSOLA))
    print("-" * ANCHO_CONSOLA)
    print("Universidad Peruana de Ciencias Aplicadas".center(ANCHO_CONSOLA))
    if logo_ascii:
        print(" " * ANCHO_CONSOLA)
        for linea in logo_ascii: print(ROJO + linea.center(ANCHO_CONSOLA) + RESET)
        print(" " * ANCHO_CONSOLA)
    print("-" * ANCHO_CONSOLA)
    print(f"{ROJO}{NEGRITA}{'EXÍGETE, INNOVA'.center(ANCHO_CONSOLA)}{RESET}")
    print(f"{ROJO}{NEGRITA}{'UPC'.center(ANCHO_CONSOLA)}{RESET}")
    print("=" * ANCHO_CONSOLA + "\n")

def mostrar_menu():
    print(f"{NEGRITA}--- MENÚ PRINCIPAL DE MONITOREO ---{RESET}")
    print(f"{VERDE}[1]{RESET} Detección de ARP Spoofing")
    print(f"{VERDE}[2]{RESET} Monitor de Ancho de Banda")
    print(f"{VERDE}[3]{RESET} Herramienta Ping Manual (Diagnóstico)")
    print(f"{VERDE}[4]{RESET} Monitor de Disponibilidad 24/7 (Tiempo Real)")
    print(f"{VERDE}[5]{RESET} Escáner de Red / Lista Blanca")
    print(f"{VERDE}[6]{RESET} Ver reporte de logs y Excel")
    print(f"{ROJO}[7] Salir{RESET}")
    print("-" * 40)

def enviar_email_con_adjuntos(asunto, cuerpo_texto, archivos):
    """
    Envía un correo con archivos adjuntos (Excel) + firma + logo.
    """
    try:
        msg = EmailMessage()
        msg['Subject'] = asunto
        msg['From'] = EMAIL_USER
        msg['To'] = EMAIL_DESTINO

        # Texto plano
        msg.set_content(cuerpo_texto)

        # HTML con firma
        texto_html = cuerpo_texto.replace('\n', '<br>')
        firma_html = """
        <br><br>
        <strong>Saludos cordiales</strong><br>
        <strong>Grupo 3</strong><br><br>
        <img src="cid:logo_upc" alt="Logo Grupo 3" width="150">
        """

        html_content = f"""
        <html>
            <body style="font-family: Arial, sans-serif; color: #333;">
                <p>{texto_html}{firma_html}</p>
            </body>
        </html>
        """
        msg.add_alternative(html_content, subtype='html')

        # Adjuntar logo
        try:
            with open(NOMBRE_LOGO, 'rb') as img:
                img_data = img.read()
                maintype, subtype = mimetypes.guess_type(NOMBRE_LOGO)[0].split('/')
                msg.get_payload()[1].add_related(img_data, maintype=maintype, subtype=subtype, cid='logo_upc')
        except: pass

        # Adjuntar archivos (Excel)
        for archivo in archivos:
            with open(archivo, 'rb') as f:
                data = f.read()
                nombre = os.path.basename(archivo)
                msg.add_attachment(data, maintype='application', subtype='octet-stream', filename=nombre)

        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASS)
        server.send_message(msg)
        server.quit()

        print(">> [EXITO] Correo con reporte enviado correctamente.")
        return True

    except Exception as e:
        print(f">> [ERROR EMAIL ADJUNTO] {e}")
        return False

# ==============================================================================
# 4. MÓDULO ARP SPOOFING
# ==============================================================================
def enviar_alerta_arp(ip, mac_falsa, mac_real):
    """
    Envía correo de alerta ARP con formato HTML, firma y logo.
    """
    mi_ip = obtener_ip_local()

    cuerpo = (
        f"<strong>ALERTA DE SEGURIDAD ARP</strong>\n\n"
        f"Reportado por el equipo: {mi_ip}\n"
        f"-----------------------------------\n"
        f"IP Objetivo: {ip}\n"
        f"MAC Intruso: <span style='color:red; font-weight:bold'>{mac_falsa}</span>\n"
        f"MAC Real: <span style='color:green; font-weight:bold'>{mac_real}</span>\n"
        f"-----------------------------------\n"
        f"Trabajo Final Fundamentos de Programación\n"
        f"Grupo 3\n"
        f"Universidad Peruana de Ciencias Aplicadas\n"
        f"<strong>Exígete, Innova, UPC</strong>"
    )

    print(" >> [ENVIANDO] Generando reporte ARP...")
    enviar_email_con_adjuntos(f"[ALERTA URGENTE] ARP Spoofing - {ip}", cuerpo, [])
    print(" >> [EXITO] Alerta ARP enviada correctamente.")
   

def analizar_paquete_arp(pkt):
    global ultima_alerta_arp
    if pkt.haslayer(ARP) and pkt[ARP].op == 2: 
        ip_src = pkt[ARP].psrc
        mac_src = pkt[ARP].hwsrc
        
        if ip_src in IP_MAC_CONFIANZA:
            mac_real = IP_MAC_CONFIANZA[ip_src]
            if mac_src != mac_real:
                ahora = time.time()
                if (ahora - ultima_alerta_arp) > TIEMPO_ENTRE_ALERTAS:
                    aviso = f"\n{ROJO}[!!!] ATAQUE ARP: IP {ip_src} suplantada por {mac_src}{RESET}"
                    print(aviso)
                    logging.warning(aviso)
                    enviar_alerta_arp(ip_src, mac_src, mac_real)
                    ultima_alerta_arp = ahora

def ejecutar_arp_sniffer():
    print("=================================================")
    print(f"{AMARILLO}[*] MONITOR ARP ACTIVADO (Ctrl + C para salir){RESET}")
    print(f"[*] Mi IP Local: {obtener_ip_local()}")
    print("=================================================")
    try:
        sniff(iface=NOMBRE_INTERFAZ, filter="arp", prn=analizar_paquete_arp, store=0)
    except KeyboardInterrupt:
        print(f"\n{VERDE}[*] Deteniendo ARP...{RESET}")
        time.sleep(1)
        return

# ==============================================================================
# 5. MÓDULO ANCHO DE BANDA
# ==============================================================================
def generar_grafico_trafico():
    print(f"[*] Generando gráfico optimizado...")
    try:
        df = pd.read_csv(ARCHIVO_CSV)
        if df.empty: return

        df['Bajada_MB'] = df['Bytes_Bajada'] / (1024 * 1024)
        df['Subida_MB'] = df['Bytes_Subida'] / (1024 * 1024)
        
        plt.figure(figsize=(12, 6))
        plt.plot(df.index, df['Bajada_MB'], label='Bajada (MB/s)', color='green')
        plt.plot(df.index, df['Subida_MB'], label='Subida (MB/s)', color='blue', linestyle='--')
        
        total = len(df)
        if total > 10:
            salto = total // 10
            idxs = range(0, total, salto)
            plt.xticks(idxs, df['Tiempo'].iloc[idxs], rotation=45)
        else:
            plt.xticks(range(total), df['Tiempo'], rotation=45)

        plt.title('Consumo de Ancho de Banda')
        plt.legend(); plt.grid(True); plt.tight_layout()
        plt.savefig(ARCHIVO_IMG); plt.close()
        print(f"[*] Gráfico guardado: {ARCHIVO_IMG}")
    except Exception as e: print(f"[!] Error gráfico: {e}")

def monitor_ancho_banda():
    global ultima_alerta_bw
    mi_ip = obtener_ip_local()
    print("=================================================")
    print(f"{AMARILLO}[*] MONITOR DE ANCHO DE BANDA ACTIVADO (Ctrl + C para salir){RESET}")
    print(f"[*] Alerta > {UMBRAL_MBPS} Mbps")
    print(f"[*] IP Monitor: {mi_ip}")
    print("=================================================")

    with open(ARCHIVO_CSV, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Tiempo", "Bytes_Bajada", "Bytes_Subida"])

    contadores_antiguos = psutil.net_io_counters(pernic=True)

    try:
        while True:
            time.sleep(1)
            contadores_nuevos = psutil.net_io_counters(pernic=True)
            if NOMBRE_INTERFAZ in contadores_nuevos:
                datos_nuevos = contadores_nuevos[NOMBRE_INTERFAZ]
                datos_antiguos = contadores_antiguos[NOMBRE_INTERFAZ]

                b_recv = datos_nuevos.bytes_recv - datos_antiguos.bytes_recv
                b_sent = datos_nuevos.bytes_sent - datos_antiguos.bytes_sent

                mbps_bajada = (b_recv * 8) / 1_000_000
                mbps_subida = (b_sent * 8) / 1_000_000
                total_mbps = mbps_bajada + mbps_subida
                hora = datetime.datetime.now().strftime("%H:%M:%S")

                col_b = ROJO if mbps_bajada > UMBRAL_MBPS else VERDE
                col_s = ROJO if mbps_subida > UMBRAL_MBPS else AZUL
                
                print(f"[{hora}] Bajada: {col_b}{convertir_bytes(b_recv)}/s{RESET} | Subida: {col_s}{convertir_bytes(b_sent)}/s{RESET}")

                if total_mbps > UMBRAL_MBPS:
                    ahora = time.time()
                    if (ahora - ultima_alerta_bw) > TIEMPO_ENTRE_ALERTAS_BW:
                        print(f" {ROJO}[!] ALERTA ENVIADA AL CORREO{RESET}")
                        logging.warning(f"ALERTA TRÁFICO: {total_mbps:.2f} Mbps")
                        
                        cuerpo = (f"<strong>ALERTA DE ANCHO DE BANDA CRÍTICO</strong>\n\n"
                                  f"Reportado por el equipo: {mi_ip}\n"
                                  f"Hora del evento: {hora}\n"
                                  f"-----------------------------------\n"
                                  f"Velocidad Total: <span style='color:red; font-weight:bold'>{total_mbps:.2f} Mbps</span>\n"
                                  f"Descarga: {mbps_bajada:.2f} Mbps\n"
                                  f"Subida: {mbps_subida:.2f} Mbps\n"
                                  f"-----------------------------------\n"
                                  f"Trabajo Final Fundamentos de Programacion\n"
                                  f"Grupo 3\n"
                                  f"Universidad Peruana de Ciencias Aplicadas\n"
                                  f"<strong>Exigete, Innova, UPC</strong>")
                        
                        enviar_email_con_adjuntos(f"[ALERTA TRÁFICO CRÍTICO] {total_mbps:.0f} Mbps Detectados", cuerpo, [])
                        ultima_alerta_bw = ahora

                with open(ARCHIVO_CSV, mode='a', newline='') as file:
                    writer = csv.writer(file)
                    writer.writerow([hora, b_recv, b_sent])
                contadores_antiguos = contadores_nuevos
            else:
                print(f"Error: Interfaz {NOMBRE_INTERFAZ} no encontrada."); break
            
    except KeyboardInterrupt:
        print("\n[*] Deteniendo monitor...")
        if input("¿Generar gráfico? (s/n): ").lower() == 's': generar_grafico_trafico()
        return

# ==============================================================================
# 6. MÓDULO PING Y DISPONIBILIDAD
# ==============================================================================

def ping_cargar_servidores():
    servidores = []
    try:
        with open(PING_ARCHIVO_SERVIDORES, "r") as archivo:
            lector = csv.reader(archivo)
            for fila in lector:
                if len(fila) >= 2:
                    nombre = fila[0].strip(); ip = fila[1].strip()
                    if nombre and ip: servidores.append((nombre, ip))
    except IOError:
        servidores = PING_SERVIDORES_DEFECTO[:]
    return servidores

def ping_hacer_ping(ip_destino, timeout_ms=1000):
    sistema = platform.system().lower()
    if sistema == "windows":
        comando = f"ping -n 1 -w {timeout_ms} {ip_destino} > nul"
    else:
        timeout_seg = max(1, int(timeout_ms / 1000))
        comando = f"ping -c 1 -W {timeout_seg} {ip_destino} > /dev/null 2>&1"
    
    t1 = time.time()
    resultado = os.system(comando)
    t2 = time.time()
    if resultado == 0: return True, (t2 - t1) * 1000.0
    else: return False, None

def ping_guardar_historial(mediciones):
    existe = os.path.exists(PING_ARCHIVO_HISTORIAL)
    with open(PING_ARCHIVO_HISTORIAL, "a", newline="") as archivo:
        escritor = csv.writer(archivo)
        if not existe:
            escritor.writerow(["nombre", "ip", "intento", "timestamp", "exito", "rtt_ms"])
        for m in mediciones: escritor.writerow(m)

def ping_resumen_por_servidor(mediciones):
    resumen = {}
    for nombre, ip, _, _, exito, rtt_ms in mediciones:
        clave = (nombre, ip)
        if clave not in resumen:
            resumen[clave] = {"total":0, "exitos":0, "fallos":0, "suma_rtt":0.0}
        resumen[clave]["total"] += 1
        if exito:
            resumen[clave]["exitos"] += 1
            resumen[clave]["suma_rtt"] += rtt_ms
        else:
            resumen[clave]["fallos"] += 1
    return resumen

def ping_guardar_resumen_csv(resumen):
    with open(PING_ARCHIVO_RESUMEN, "w", newline="", encoding="utf-8") as archivo:
        writer = csv.writer(archivo)
        writer.writerow(["Servidor", "IP", "Total_Pings", "Exitos", "Fallos", "Disponibilidad_%", "RTT_Promedio_ms"])
        for (nombre, ip), datos in resumen.items():
            total = datos["total"]
            exitos = datos["exitos"]
            fallos = datos["fallos"]
            disp = (exitos * 100 / total) if total > 0 else 0
            rtt_prom = (datos["suma_rtt"] / exitos) if exitos > 0 else 0
            writer.writerow([nombre, ip, total, exitos, fallos, round(disp, 2), round(rtt_prom, 2)])

def ping_clasificar_estado(disponibilidad):
    if disponibilidad >= PING_UMBRAL_EXCELENTE: return f"{VERDE}EXCELENTE{RESET}"
    elif disponibilidad >= PING_UMBRAL_BUENO: return f"{VERDE}BUENA{RESET}"
    elif disponibilidad >= PING_UMBRAL_ACEPTABLE: return f"{AMARILLO}ACEPTABLE{RESET}"
    else: return f"{ROJO}CRÍTICA{RESET}"

def enviar_alerta_ping(nombre_servidor, ip_servidor):
    mi_ip = obtener_ip_local()
    cuerpo = (f"<strong>ALERTA DE DISPONIBILIDAD DE SERVICIO</strong>\n\n"
              f"Reportado por el monitor: {mi_ip}\n"
              f"-----------------------------------\n"
              f"Se ha detectado una caída de servicio crítica.\n\n"
              f"Servidor: <strong>{nombre_servidor}</strong>\n"
              f"IP Objetivo: <span style='color:red; font-weight:bold'>{ip_servidor}</span>\n"
              f"Estado: <span style='background-color:red; color:white; padding:2px'>OFFLINE / NO RESPONDE</span>\n"
              f"-----------------------------------\n"
              f"Acción sugerida: Verificar conectividad física o reiniciar el servicio.\n"
              f"Trabajo Final Fundamentos de Programacion\n"
              f"Grupo 3 - UPC")
    
    print(f" >> [ENVIANDO] Alerta crítica por correo a los administradores...")
    enviar_email_con_adjuntos(f"[ALERTA CRÍTICA] Servidor Caído - {nombre_servidor}", cuerpo, [])

# --- FUNCIONES PRINCIPALES DE PING ---

def ejecutar_ping_automatico():
    """
    Opción 3: HERRAMIENTA MANUAL DE DIAGNÓSTICO
    """
    limpiar_pantalla()
    print(f"{NEGRITA}--- HERRAMIENTA DE DIAGNÓSTICO (PING MANUAL) ---{RESET}\n")

    target = input(f"Ingrese la IP o Dominio a probar (Ej: 8.8.8.8): {AMARILLO}")
    print(f"{RESET}", end="") 
    
    if not target:
        print(f"{ROJO}Error: No ingresaste ninguna IP.{RESET}")
        input("Presiona Enter para volver..."); return

    try:
        veces_input = input(f"¿Cuántas repeticiones? [Defecto: 4]: {AMARILLO}")
        print(f"{RESET}", end="")
        veces = int(veces_input) if veces_input else 4
    except ValueError:
        veces = 4

    print(f"\n{AZUL}Haciendo ping a {NEGRITA}{target}{RESET}{AZUL} con {veces} paquetes de datos:{RESET}\n")

    enviados = 0
    recibidos = 0
    tiempos = []

    try:
        for i in range(veces):
            enviados += 1
            exito, rtt = ping_hacer_ping(target, timeout_ms=1000)

            if exito:
                recibidos += 1
                tiempos.append(rtt)
                print(f"Respuesta desde {target}: tiempo={VERDE}{int(rtt)}ms{RESET}")
            else:
                print(f"{ROJO}Tiempo de espera agotado para esta solicitud.{RESET}")
            time.sleep(1)

        perdidos = enviados - recibidos
        porcentaje_perdida = (perdidos / enviados) * 100

        print(f"\n{NEGRITA}--- Estadísticas de ping para {target} ---{RESET}")
        print(f"    Paquetes: enviados = {enviados}, recibidos = {recibidos}, perdidos = {perdidos} ({int(porcentaje_perdida)}% perdidos)")

        if tiempos:
            minimo = min(tiempos)
            maximo = max(tiempos)
            promedio = sum(tiempos) / len(tiempos)
            print(f"Tiempos aproximados de ida y vuelta en milisegundos:")
            print(f"    Mínimo = {int(minimo)}ms, Máximo = {int(maximo)}ms, Media = {AZUL}{int(promedio)}ms{RESET}")

    except KeyboardInterrupt:
        print(f"\n\n{AMARILLO}[!] Ping cancelado por usuario.{RESET}")
    
    input("\nPresiona Enter para volver al menú...")


def ejecutar_monitoreo_disponibilidad():
    """
    Opción 4: MONITOR INTELIGENTE (BUCLE INFINITO)
    """
    limpiar_pantalla()
    print(f"{NEGRITA}--- MONITOR DE DISPONIBILIDAD (VIGILANCIA 24/7) ---{RESET}")
    print(f"{AMARILLO}[*] Presiona Ctrl + C para detener y volver al menú.{RESET}")
    print(f"[*] Iniciando vigilancia de servidores...\n")
    
    servidores = ping_cargar_servidores()
    if not servidores:
        print(f"{ROJO}No hay servidores configurados.{RESET}")
        input("Enter para volver..."); return

    estado_previo = {}

    try:
        while True:
            for nombre, ip in servidores:
                exito, rtt = ping_hacer_ping(ip, timeout_ms=1000)
                
                if exito: estado_actual = "ONLINE"
                else: estado_actual = "OFFLINE"

                if ip not in estado_previo:
                    color = VERDE if estado_actual == "ONLINE" else ROJO
                    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] Estado Inicial: {nombre} ({ip}) -> {color}{estado_actual}{RESET}")
                    estado_previo[ip] = estado_actual
                
                elif estado_previo[ip] != estado_actual:
                    hora = datetime.datetime.now().strftime('%H:%M:%S')
                    if estado_actual == "OFFLINE":
                        print(f"{ROJO}[{hora}] ¡ALERTA! {nombre} ({ip}) ha cambiado a OFFLINE{RESET}")
                        logging.warning(f"MONITOR: {nombre} ({ip}) CAÍDO (OFFLINE)")
                        enviar_alerta_ping(nombre, ip)
                    else:
                        print(f"{VERDE}[{hora}] RESTAURADO: {nombre} ({ip}) ha vuelto a ONLINE{RESET}")
                        logging.warning(f"MONITOR: {nombre} ({ip}) RECUPERADO (ONLINE)")
                    estado_previo[ip] = estado_actual
            
            time.sleep(2)

    except KeyboardInterrupt:
        print(f"\n\n{AMARILLO}[!] Deteniendo monitor... Volviendo al menú principal.{RESET}")
        time.sleep(1)
        return

# ==============================================================================
# 7. MÓDULO ESCÁNER DE RED
# ==============================================================================
def enviar_alerta_desconocido(lista_desconocidos):
    mi_ip = obtener_ip_local()
    items_html = ""
    for ip, mac in lista_desconocidos:
        items_html += f"<li>IP: <strong>{ip}</strong> - MAC: {mac}</li>"

    cuerpo = (f"<strong>ALERTA DE SEGURIDAD - DISPOSITIVOS NO AUTORIZADOS</strong>\n\n"
              f"Reportado por: {mi_ip}\n"
              f"-----------------------------------\n"
              f"El escáner de red ha detectado dispositivos que NO están en la Lista Blanca:\n"
              f"<ul><span style='color:red; font-weight:bold'>{items_html}</span></ul>\n"
              f"-----------------------------------\n"
              f"Acción sugerida: Verificar si son invitados o intrusos.\n"
              f"-----------------------------------\n"
              f"Trabajo Final Fundamentos de Programacion\n"
              f"Grupo 3\n"
              f"Universidad Peruana de Ciencias Aplicadas\n"
              f"<strong>Exigete, Innova, UPC</strong>")
    
    print(f" >> [ENVIANDO] Alerta de dispositivos desconocidos")
    enviar_email_con_adjuntos("[ALERTA] Intrusos en la Red Detectados", cuerpo, [])
    print(f" >> [EXITO] Informe enviado al administrador.")

def ejecutar_scanner():
    limpiar_pantalla()
    print(f"{NEGRITA}--- MÓDULO: Escáner de Dispositivos (Lista Blanca) ---{RESET}")
    print(f"[*] Escaneando rango: {AMARILLO}{RANGO_RED_SCANNER}{RESET}")
    print("[*] Por favor espere, esto puede tardar unos segundos...\n")

    arp = ARP(pdst=RANGO_RED_SCANNER)
    ether = Ether(dst="ff:ff:ff:ff:ff:ff")
    packet = ether/arp

    try:
        result = srp(packet, timeout=2, verbose=0, iface=NOMBRE_INTERFAZ)[0]
    except Exception as e:
        print(f"{ROJO}[ERROR] Falló el escaneo: {e}{RESET}")
        print("Verifique que el nombre de la interfaz sea correcto.")
        input("Enter para volver..."); return

    unknown_devices = []
    print(f"{'IP':<16} {'MAC':<20} {'ESTADO'}")
    print("-" * 60)

    for sent, received in result:
        ip = received.psrc
        mac = received.hwsrc
        
        if ip in IP_MAC_CONFIANZA and IP_MAC_CONFIANZA[ip] == mac:
            estado = f"{VERDE}AUTORIZADO{RESET}"
        elif ip in IP_MAC_CONFIANZA and IP_MAC_CONFIANZA[ip] != mac:
            estado = f"{ROJO}CONFLICTO MAC{RESET}"
            unknown_devices.append((ip, mac))
        else:
            estado = f"{AMARILLO}DESCONOCIDO{RESET}"
            unknown_devices.append((ip, mac))

        print(f"{ip:<16} {mac:<20} {estado}")

    print("-" * 60)

    if unknown_devices:
        print(f"\n{ROJO}[!] Se detectaron {len(unknown_devices)} dispositivos sospechosos.{RESET}")
        for ip, mac in unknown_devices:
            logging.warning(f"INTRUSO DETECTADO: IP {ip} MAC {mac} no está en lista blanca.")
        
        print("Generando reporte por correo...")
        enviar_alerta_desconocido(unknown_devices)
    else:
        print(f"\n{VERDE}[OK] Todos los dispositivos encontrados son confiables.{RESET}")

    input("\nPresiona Enter para volver al menú...")
    return

# ==============================================================================
# 8. FUNCIÓN PRINCIPAL
# ==============================================================================

def generar_reporte_consolidado():
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill

    eventos = []

    # LOG PRINCIPAL
    if os.path.exists("seguridad_sistema.log"):
        with open("seguridad_sistema.log", "r", encoding="utf-8", errors="ignore") as f:
            for linea in f:
                try:
                    fecha_hora, detalle = linea.strip().split(" - ", 1)
                    fecha, hora = fecha_hora.split(" ")
                    eventos.append([fecha, hora, "SEGURIDAD", detalle])
                except: pass

    # HISTORIAL ANCHO DE BANDA
    if os.path.exists(ARCHIVO_CSV):
        df_bw = pd.read_csv(ARCHIVO_CSV)
        hoy = datetime.date.today().isoformat()
        for _, r in df_bw.iterrows():
            # Compresion de lista para evitar errores de sintaxis al copiar
            info_trafico = f"Bajada:{r['Bytes_Bajada']} Subida:{r['Bytes_Subida']}"
            eventos.append([hoy, r['Tiempo'], "TRAFICO", info_trafico])

    if not eventos:
        print("No existen eventos reales para consolidar.")
        return

    # CREAR EXCEL
    df = pd.DataFrame(eventos, columns=["Fecha", "Hora", "Tipo_Alerta", "Detalle"])
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_file = f"reporte_consolidado_{timestamp}.xlsx"

    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Eventos", index=False)
        resumen = df['Tipo_Alerta'].value_counts().reset_index()
        resumen.columns = ["Tipo_Alerta", "Cantidad"]
        resumen.to_excel(writer, sheet_name="Resumen", index=False)
        fechas = df['Fecha'].value_counts().sort_index().reset_index()
        fechas.columns = ["Fecha", "Cantidad"]
        fechas.to_excel(writer, sheet_name="Eventos_por_Fecha", index=False)

    wb = load_workbook(excel_file)
    ws_res = wb["Resumen"]
    ws_fec = wb["Eventos_por_Fecha"]
    ws_dash = wb.create_sheet("Dashboard")

    fondo = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    for row in ws_dash.iter_rows(min_row=1, max_row=40, min_col=1, max_col=20):
        for cell in row:
            cell.fill = fondo

    ws_dash.merge_cells("A1:J2")
    ws_dash["A1"] = "REPORTE DE INDICADORES DE SEGURIDAD – GRUPO 3"
    ws_dash["A1"].font = Font(size=20, bold=True)
    ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")

    bar = BarChart()
    bar.title = "Alertas por Tipo"
    data = Reference(ws_res, min_col=2, min_row=1, max_row=ws_res.max_row)
    cats = Reference(ws_res, min_col=1, min_row=2, max_row=ws_res.max_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws_dash.add_chart(bar, "B4")

    line = LineChart()
    line.title = "Eventos en el Tiempo"
    data = Reference(ws_fec, min_col=2, min_row=1, max_row=ws_fec.max_row)
    cats = Reference(ws_fec, min_col=1, min_row=2, max_row=ws_fec.max_row)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    ws_dash.add_chart(line, "B20")

    wb.save(excel_file)
    os.startfile(excel_file)

    enviar_email_con_adjuntos(
        "REPORTE DE INDICADORES DE SEGURIDAD – GRUPO 3",
        "Reporte generado únicamente con eventos reales del sistema.",
        [excel_file]
    )

def main():
    try:
        crear_caratula(NOMBRE_LOGO)
        input(f"{AMARILLO}Presione ENTER para continuar al menú principal...{RESET}")
        limpiar_pantalla()

        while True:
            limpiar_pantalla()
            mostrar_menu()
            opcion = input("Opción: ").strip()

            if opcion == "1":
                ejecutar_arp_sniffer()
            elif opcion == "2":
                monitor_ancho_banda()
            elif opcion == "3":
                ejecutar_ping_automatico()
            elif opcion == "4":
                ejecutar_monitoreo_disponibilidad()
            elif opcion == "5":
                ejecutar_scanner()
            elif opcion == "6":
                generar_reporte_consolidado()
            elif opcion == "7":
                print("Saliendo del sistema...")
                break
            else:
                print("Opción inválida")
                time.sleep(1)

    except KeyboardInterrupt:
        print(f"\n{AMARILLO}Ejecución interrumpida por el usuario. Saliendo del sistema...{RESET}")

if __name__ == "__main__":
    main()