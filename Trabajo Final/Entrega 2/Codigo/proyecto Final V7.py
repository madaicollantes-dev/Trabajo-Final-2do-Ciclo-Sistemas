# ==============================================================================
# 1. IMPORTACIÓN DE LIBRERÍAS
# ==============================================================================
import logging                           # Para guardar registro de eventos (logs)
import smtplib                           # Para enviar correos electrónicos
import time                              # Para manejar tiempos y pausas
import os                                # Para limpiar pantalla del sistema
import psutil                            # Para leer el tráfico de la tarjeta de red
import csv                               # Para leer y escribir archivos en formato csv
import datetime                          # Para manejar y realizar operaciones con el tiempo fechas, horas, minutos y segundos
import socket                            # Para acceder a la tarjeta de red y obtener la ip local
import pandas as pd                      # Para analisis y manipulacion, lee y organiza inteligentemente los datos para el grafico
import mimetypes                         # Para identificar el tipo de imagen del logo y adjuntarlo al correo
import matplotlib
matplotlib.use("Agg")                    # Backend sin interfaz gráfica (evita Tkinter)
import matplotlib.pyplot as plt          # Para creacion de graficos
import platform
from email.message import EmailMessage   # Para dar formato al correo
from scapy.all import sniff, ARP, Ether, srp, conf   # Scapy: El núcleo del análisis de red
from PIL import Image                    # Pillow: Para procesar la imagen del logo
# ==============================================================================
# 2. CONFIGURACIONES Y COLORES
# ==============================================================================
ROJO = '\033[91m'             # Secuencias de escape ANSI todos inician con \033[
VERDE = '\033[92m'            # \033: Indica que no es un texto para imprimir, si no una instruccion que seguir 
AMARILLO = '\033[93m'         # [: inicia la instruccion
AZUL = '\033[94m'             # numero XX:en este caso el 94 indica color azul y m indica negrita 
RESET = '\033[0m'             # en este caso el 0m apaga todos los efectos
NEGRITA = '\033[1m'
ANCHO_CONSOLA = 60            # Ancho consola sirve para definir un tamaño de referencia y poder definir un centro

# --- CONFIGURACIÓN DE CORREO ---
SMTP_SERVER = "smtp.gmail.com"               #Si usas google como correo de envio dejalo asi
SMTP_PORT = 587                              #Si usas google como correo de envio dejalo asi
EMAIL_USER = "coblitasv2@gmail.com"          #Colocar el correo desde el que enviaras las alertas
EMAIL_PASS = "nyaj rouc jenj usuj"           #Colocar la contraseña de la API (no la del correo)
EMAIL_DESTINO = "u202514590@upc.edu.pe"      #Colocar el correo del profe
NOMBRE_LOGO = "descarga.png"                 #La imagen que se usará en la firma y logo de caratula

# --- CONFIGURACIÓN DE RED ---
IP_MAC_CONFIANZA = {
    "192.168.1.1" : "74:93:da:e8:44:80",  #Reemplazar por la ip y mac de su router o gateway
    "192.168.1.42": "a8:3b:76:19:07:cd",  #Reemplazar por la ip y mac de su laptop o PC desde la que ejecutaran el codigo
    "192.168.1.45": "a0:b1:c2:d3:e4:f5",  #Reemplazar por la ip de un equipo cualquiera que este conectado a la red la mac puede ser cualquiera (debe ser falsa) ya que esto activiara las alertas
}
NOMBRE_INTERFAZ = "Wi-Fi" #Colocar el nombre de su interface de red Sea Wifi o Eth
#Rango de Red a Revisar (MADAI)
RANGO_RED_SCANNER = "192.168.1.0/24"
# --- CONFIGURACIÓN DE ALERTAS ---  
TIEMPO_ENTRE_ALERTAS = 10     # Define el tiempo de cooldown en segundos para alertas de ataques. Sin esto recibiría 5,000 correos en un minuto y bloquearía mi bandeja de entrada
TIEMPO_ENTRE_ALERTAS_BW = 60  # Se configura en 60 (1 minuto). Si fuera muy corto (ej. 10s), recibirías un correo nuevo cada 10 segundos 
UMBRAL_MBPS = 100             # Es el límite de velocidad (en Megabits por segundo) que consideramos "peligroso" o "anormal"
# --- Integración: Raphael Avalos ---
PING_ARCHIVO_SERVIDORES = "servidores.csv"
PING_ARCHIVO_HISTORIAL = "historial_ping.csv"
PING_ARCHIVO_RESUMEN = "ping_resumen.csv"
PING_SERVIDORES_DEFECTO = [
    ("Servidor AD",   "192.168.1.42"), #Reemplazar por la ip y mac de su laptop o PC desde la que ejecutaran el codigo
    ("Servidor WEB",  "192.168.1.45"), #Reemplazar por la ip de un equipo cualquiera que este conectado a la red la mac puede ser cualquiera (debe ser falsa) ya que esto activiara las alertas
    ("Puerta de Enlace",  "192.168.1.1"), #Reemplazar por la ip de su router.
    ("Servidor DNS",  "8.8.8.8")
]
PING_UMBRAL_EXCELENTE = 99.0
PING_UMBRAL_BUENO      = 95.0
PING_UMBRAL_ACEPTABLE  = 80.0
# Variables de estado
ultima_alerta_arp = 0         # Esas variables se inicializan en 0 para representar "El inicio de los tiempos" 
ultima_alerta_bw = 0          # Nunca he enviado una alerta antes

#logging = libreria de registro de eventos
logging.basicConfig(filename='seguridad_sistema.log', level=logging.WARNING, 
                    format='%(asctime)s - %(message)s')
#filename='seguridad_sistema.log': Le dice al programa: "No imprimas los errores solo en la pantalla; guárdalos en este archivo de texto". Si el archivo no existe, lo crea
#level=level=logging.WARNING: Es el filtro de importancia. Le dice al sistema: "Ignora los mensajes informativos (INFO) o de depuración (DEBUG). Solo guarda cosas importantes como Advertencias (WARNING)
#format='%(asctime)s - %(message)s': Define cómo se ve cada línea en el archivo de texto

#Estas dos líneas configuran conf, que es la configuración global de la librería Scapy
conf.verb = 0       # Modo silencioso para scapy. Evita que ensucie la interface grafica con spam de anuncios de monitoreo
conf.resolve = []   # Desactivamos el DNS para agilizar el escaneo

#Estas dos lineas definen el nombre y formato de los archivos que generaremos al monitorear el ancho de banda
ARCHIVO_CSV = 'historial_trafico.csv'
ARCHIVO_IMG = 'grafico_trafico.png'

# ==============================================================================
# 3. FUNCIONES AUXILIARES (UI y NET)
# ==============================================================================
def limpiar_pantalla():
    # Pregunta: "¿Estoy en Windows ('nt')?"
    if os.name == 'nt': os.system('cls') # Si es sí, usa el comando de limpieza de Windows.
    else: os.system('clear') # Si es no (Mac o Linux), usa el comando de ellos.

def convertir_bytes(peso):
    # Revisa cada "byte" disponible: Bytes, KiloBytes, MegaBytes, GigaBytes
    for unidad in ['B', 'KB', 'MB', 'GB']:
        # Si el tamaño es pequeño (menor a 1024), se queda con esa unidad.
        if peso < 1024.0:                
            return f"{peso:.2f} {unidad}" # Devuelve el número con 2 decimales (ej. 10.50 MB)
        # Si es muy grande, lo divide entre 1024 para pasar a la siguiente unidad.
        # Ejemplo: 2048 Bytes / 1024 = 2 KB
        peso /= 1024.0

def obtener_ip_local():
    """Obtiene la IP real de la máquina."""
    # Crea un socket para conectarse a internet.
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        # Intenta conectar con el DNS de Google (8.8.8.8).
        # Nota: No envía datos reales, solo establece la intención de ruta.
        s.connect(('8.8.8.8', 1))
        # Le pregunta al sistema: "¿Qué IP usaste para intentar esta conexión?"
        IP = s.getsockname()[0]
    except Exception:
        # Si no tienes internet o falla, asume que eres "localhost" (127.0.0.1)
        IP = '127.0.0.1' #Esta ip es del localhost no debe modificarse 
    finally:
        # Cierra el socket para no gastar recursos.
        s.close()
    return IP

def crear_caratula(ruta_imagen, ancho_logo=30):
    limpiar_pantalla()  # Borra la pantalla primero
    logo_ascii = []     # Lista vacía donde guardaremos el logo
    try:
        img = Image.open(ruta_imagen)  # Abre la imagen desde el archivo
        # --- Matemáticas de proporción ---
        # Calcula qué tan alta es la imagen comparada con su ancho
        ancho_orig, alto_orig = img.size
        relacion = alto_orig / ancho_orig
        # Calcula la nueva altura. Se multiplica por 0.55 porque las letras en 
        # la consola son rectangulares (altas), no cuadradas. Esto evita que el logo se vea estirado
        alto_logo = int(ancho_logo * relacion * 0.55)
        img = img.resize((ancho_logo, alto_logo)).convert("L")              # Redimensiona la imagen y la convierte a Blanco y Negro ("L")
        pixeles = img.getdata()                        # Obtiene todos los puntos (píxeles) de la imagen
        cadena_ascii = "".join([" " if p > 150 else "█" for p in pixeles])  # Si el punto es blanco (>150), pon un espacio. Si es oscuro, pon un bloque "█"
        logo_ascii = [cadena_ascii[i:i+ancho_logo] for i in range(0, len(cadena_ascii), ancho_logo)] # Corta la cadena larga en líneas para formar el cuadrado del dibujo
    except Exception: pass             # Si la imagen falla, no hace nada y sigue (no rompe el programa)

    # --- Impresión del Texto ---
    # .center() centra el texto en la pantalla basándose en el ancho definido
    print("\n" + "=" * ANCHO_CONSOLA)
    print(f"{NEGRITA}{'TRABAJO FINAL: FUNDAMENTOS DE PROGRAMACIÓN'.center(ANCHO_CONSOLA)}{RESET}")
    print("-" * ANCHO_CONSOLA)
    print("Profesor: Luis Castillo Mesias".center(ANCHO_CONSOLA))
    print("NRC:803".center(ANCHO_CONSOLA))
    print("Carrera de Ingeniería de Redes y Comunicaciones".center(ANCHO_CONSOLA))
    print(" " * ANCHO_CONSOLA)
    print("Grupo 3".center(ANCHO_CONSOLA))
    print("Integrantes: \n -Carlos Jesus Oblitas Vargas\n -Madai Gonzales Collantes\n -Raphael Enrique Avalos Espinoza\n -Bils Ortega Buitron".center(ANCHO_CONSOLA))
    print("-" * ANCHO_CONSOLA)
    print("Universidad Peruana de Ciencias Aplicadas".center(ANCHO_CONSOLA))
    if logo_ascii:   # Si logramos crear el logo, lo imprimimos línea por línea en ROJO
        print(" " * ANCHO_CONSOLA)
        for linea in logo_ascii: print(ROJO + linea.center(ANCHO_CONSOLA) + RESET)
        print(" " * ANCHO_CONSOLA)
    print("-" * ANCHO_CONSOLA)
    print(f"{ROJO}{NEGRITA}{'EXÍGETE, INNOVA'.center(ANCHO_CONSOLA)}{RESET}")
    print(f"{ROJO}{NEGRITA}{'UPC'.center(ANCHO_CONSOLA)}{RESET}")
    print("=" * ANCHO_CONSOLA + "\n")

def mostrar_menu():
    print(f"{NEGRITA}--- MENÚ PRINCIPAL DE MONITOREO ---{RESET}")
    print(f"{VERDE}[1]{RESET} Detección de ARP Spoofing")
    print(f"{VERDE}[2]{RESET} Monitor de Ancho de Banda")
    print(f"{VERDE}[3]{RESET} Herramienta Ping (Diagnóstico)") # Manual
    print(f"{VERDE}[4]{RESET} Monitor de Disponibilidad 24/7")
    print(f"{VERDE}[5]{RESET} Escáner de Red / Lista Blanca")
    print(f"{VERDE}[6]{RESET} Ver reporte de logs")
    print(f"{ROJO}[7] Salir{RESET}")
    print("-" * 40)

#--------------------------------#
# --- FUNCIÓN DE CORREO ---#
#-------------------------------#
def enviar_email_con_adjuntos(asunto, cuerpo_texto, archivos):
    """
    Envía un correo con archivos adjuntos (Excel) + firma + logo.
    """
    try:
        msg = EmailMessage()
        msg['Subject'] = asunto
        msg['From'] = EMAIL_USER
        msg['To'] = EMAIL_DESTINO

        # Texto plano
        msg.set_content(cuerpo_texto)

        # HTML con firma
        texto_html = cuerpo_texto.replace('\n', '<br>')

        firma_html = """
        <br><br>
        <strong>Saludos cordiales</strong><br>
        <strong>Grupo 3</strong><br><br>
        <img src="cid:logo_upc" alt="Logo Grupo 3" width="150">
        """

        html_content = f"""
        <html>
            <body style="font-family: Arial, sans-serif; color: #333;">
                <p>{texto_html}{firma_html}</p>
            </body>
        </html>
        """

        msg.add_alternative(html_content, subtype='html')

        # Adjuntar logo
        try:
            with open(NOMBRE_LOGO, 'rb') as img:
                img_data = img.read()
                maintype, subtype = mimetypes.guess_type(NOMBRE_LOGO)[0].split('/')
                msg.get_payload()[1].add_related(
                    img_data,
                    maintype=maintype,
                    subtype=subtype,
                    cid='logo_upc'
                )
        except:
            pass

        # Adjuntar archivos (Excel)
        for archivo in archivos:
            with open(archivo, 'rb') as f:
                data = f.read()
                nombre = os.path.basename(archivo)
                msg.add_attachment(
                    data,
                    maintype='application',
                    subtype='octet-stream',
                    filename=nombre
                )

        server = smtplib.SMTP(SMTP_SERVER, SMTP_PORT)
        server.starttls()
        server.login(EMAIL_USER, EMAIL_PASS)
        server.send_message(msg)
        server.quit()

        print(">> [EXITO] Correo con reporte enviado correctamente.")
        return True

    except Exception as e:
        print(f">> [ERROR EMAIL ADJUNTO] {e}")
        return False

# ==============================================================================
# 4. MÓDULO ARP SPOOFING
# ==============================================================================
def enviar_alerta_arp(ip, mac_falsa, mac_real):
    """
    Función encargada de redactar el correo de alerta con formato HTML.
    NO detecta el ataque, solo reporta lo que 'analizar_paquete_arp' encontró.
    """
    # Averiguamos desde qué computadora (IP) estamos vigilando para ponerlo en el informe
    mi_ip = obtener_ip_local()

    # Preparamos el cuerpo del mensaje
    # Usamos \n para los saltos de línea (la función de envío los convertirá a <br>)
    # Usamos etiquetas HTML <span> y <strong> para los colores y negritas.
    cuerpo = (f"<strong>ALERTA DE SEGURIDAD ARP</strong>\n\n"
              f"Reportado por el equipo: {mi_ip}\n"
              f"-----------------------------------\n"
              f"IP Objetivo: {ip}\n"
              f"MAC Intruso: <span style='color:red; font-weight:bold'>{mac_falsa}</span>\n"
              f"MAC Real: <span style='color:green; font-weight:bold'>{mac_real}</span>\n"
              f"-----------------------------------\n"
              f"Trabajo Final Fundamentos de Programacion\n"
              f"Grupo 3\n"
              f"Salon: 803\n"
              f"Carrera: Ingenieria de Redes y Comunicaciones\n"
              f"Universidad Peruana de Ciencias Aplicadas\n"
              f"<span style='color:black; font-weight:bold'>Exigete, Innova, UPC</span>")
    
    print(f" >> [ENVIANDO] Generando reporte ARP con firma...")
    
    # Envío del Correo:
    # Llamamos a nuestra función de correo "enviar_email", 
    # Si devuelve True, imprimimos éxito.
def enviar_alerta_arp(ip, mac_falsa, mac_real):
    """
    Envía correo de alerta ARP con formato HTML, firma y logo.
    """
    mi_ip = obtener_ip_local()

    cuerpo = (
        f"<strong>ALERTA DE SEGURIDAD ARP</strong>\n\n"
        f"Reportado por el equipo: {mi_ip}\n"
        f"-----------------------------------\n"
        f"IP Objetivo: {ip}\n"
        f"MAC Intruso: <span style='color:red; font-weight:bold'>{mac_falsa}</span>\n"
        f"MAC Real: <span style='color:green; font-weight:bold'>{mac_real}</span>\n"
        f"-----------------------------------\n"
        f"Trabajo Final Fundamentos de Programación\n"
        f"Grupo 3\n"
        f"Universidad Peruana de Ciencias Aplicadas\n"
        f"<strong>Exígete, Innova, UPC</strong>"
    )

    print(" >> [ENVIANDO] Generando reporte ARP con firma...")

    enviar_email_con_adjuntos(
        f"[ALERTA URGENTE] ARP Spoofing - {ip}",
        cuerpo,
        []
    )

    print(" >> [EXITO] Alerta ARP enviada correctamente.")
   

def analizar_paquete_arp(pkt):
    """
    Funcion de Seguridad. Se ejecuta automáticamente CADA VEZ que pasa un paquete.
    """
    global ultima_alerta_arp    # Usamos la variable global para recordar cuándo fue la ultima alerta
    # pkt.haslayer(ARP): Filtra dentro de todos los paquetes que se reciben por segundo y solo se queda con los paquetes ARP
    # pkt[ARP].op == 2: los paquetes de tipo ARP que nos interesan son el tipo "2" (respuestas) ya que el tipo "1" es de (preguntas) y los atacantes siempre responden a preguntas haciendose pasar por otra mac
    if pkt.haslayer(ARP) and pkt[ARP].op == 2: 
        ip_src = pkt[ARP].psrc      # IP de quien envió el paquete 
        mac_src = pkt[ARP].hwsrc    # MAC física de quien envió el paquete
        # Verificación en Lista Blanca:
        if ip_src in IP_MAC_CONFIANZA:
            mac_real = IP_MAC_CONFIANZA[ip_src]
            if mac_src != mac_real:
                # Control de Spam (Cooldown):
                ahora = time.time()
                if (ahora - ultima_alerta_arp) > TIEMPO_ENTRE_ALERTAS: # ¿Han pasado más de 10 segundos desde el último aviso?
                    aviso = f"\n{ROJO}[!!!] ATAQUE ARP: IP {ip_src} suplantada por {mac_src}{RESET}"
                    print(aviso)
                    logging.warning(aviso)  # Guardamos la evidencia en el log
                    # Llamada única a la función de alerta
                    enviar_alerta_arp(ip_src, mac_src, mac_real)
                    ultima_alerta_arp = ahora   # Reiniciamos el reloj para esperar otros 10 segundos antes de molestar de nuevo

def ejecutar_arp_sniffer():
    """
    Funcion que arranca la vigilancia "Sniffer"
    """
    print("=================================================")
    print(f"{AMARILLO}[*] MONITOR ARP ACTIVADO (Ctrl + C para salir){RESET}")
    print(f"[*] Mi IP Local: {obtener_ip_local()}")
    print("=================================================")
    try:
        # sniff() es la función de la librería Scapy.
        # iface: Qué tarjeta de red usar
        # filter: Solo escuchar paquetes ARP (ignora el resto para ahorrar CPU)
        # prn: "Per Packet Runtime" -> Qué función ejecutar por cada paquete
        # store=0: NO guardar paquetes en memoria RAM (evita que la PC se ponga lenta)
        sniff(iface=NOMBRE_INTERFAZ, filter="arp", prn=analizar_paquete_arp, store=0)
    except KeyboardInterrupt:
        # Si el usuario presiona Ctrl+C detenemos el Sniffer y regresamos al menu de modulos.
        print(f"\n{VERDE}[*] Deteniendo ARP...{RESET}")
        time.sleep(1)
        return

def modulo():
    print("[*] Ejecutando módulo...")

    # lógica del módulo

    input("Presiona Enter para volver al menú...")
    return


# ==============================================================================
# 5. MÓDULO ANCHO DE BANDA
# ==============================================================================
def generar_grafico_trafico():
    """
    Función que lee el historial (CSV) y crea una imagen PNG (Gráfico).
    """
    print(f"[*] Generando gráfico optimizado...")
    try:
        df = pd.read_csv(ARCHIVO_CSV)   # Usamos Pandas para abrir el 'Excel' (CSV) de forma super rápida
        if df.empty: return     # Si el archivo está vacío (no hubo datos), cancelamos para no causar error

        # Los datos vienen en 'Bytes' números grandes y dificiles de leer
        # Dividimos entre 1024*1024 para convertirlos a MegaBytes (MB), que son más legibles 
        df['Bajada_MB'] = df['Bytes_Bajada'] / (1024 * 1024)
        df['Subida_MB'] = df['Bytes_Subida'] / (1024 * 1024)
        
        plt.figure(figsize=(12, 6))   # Definimos el tamaño de la imagen (12 pulgadas de ancho x 6 de alto)

        # Dibujamos las líneas: Verde para descargas, Azul punteada para subidas
        plt.plot(df.index, df['Bajada_MB'], label='Bajada (MB/s)', color='green')
        plt.plot(df.index, df['Subida_MB'], label='Subida (MB/s)', color='blue', linestyle='--')
        
        # Limpieza del eje X
        total = len(df)     # Si mostramos TODAS las horas, se solapan y no se leen
        if total > 10:
            # Si hay muchos datos, mostramos solo 1 de cada 10 etiquetas
            salto = total // 10
            idxs = range(0, total, salto)
            plt.xticks(idxs, df['Tiempo'].iloc[idxs], rotation=45)
        else:
            # Si son pocos datos, mostramos todos.
            plt.xticks(range(total), df['Tiempo'], rotation=45)

        # Ponemos título, leyenda y guardamos la foto en el disco duro.
        plt.title('Consumo de Ancho de Banda')
        plt.legend(); plt.grid(True); plt.tight_layout()
        plt.savefig(ARCHIVO_IMG); plt.close()
        print(f"[*] Gráfico guardado: {ARCHIVO_IMG}")
    except Exception as e: print(f"[!] Error gráfico: {e}")

def monitor_ancho_banda():
    """
    Función principal que mide la velocidad segundo a segundo.
    """
    global ultima_alerta_bw     # Variable para recordar cuándo enviamos el último correo
    mi_ip = obtener_ip_local()
    # Presentación visual en consola
    print("=================================================")
    print(f"{AMARILLO}[*] MONITOR DE ANCHO DE BANDA ACTIVADO (Ctrl + C para salir){RESET}")
    print(f"[*] Alerta > {UMBRAL_MBPS} Mbps")
    print(f"[*] IP Monitor: {mi_ip}")
    print("=================================================")

    # Preparamos el archivo CSV (Borrador nuevo):
    # 'w' significa Write (Sobrescribir). Borra el historial anterior y empieza uno nuevo
    with open(ARCHIVO_CSV, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Tiempo", "Bytes_Bajada", "Bytes_Subida"])

    # TOMA 1: Leemos el contador de la tarjeta de red AHORA
    contadores_antiguos = psutil.net_io_counters(pernic=True)

    try:
        while True:
            # Pausa de 1 segundo exacto
            time.sleep(1)
            # Leemos el contador de nuevo tras 1 segundo
            contadores_nuevos = psutil.net_io_counters(pernic=True)
            # Verificamos si tu tarjeta existe
            if NOMBRE_INTERFAZ in contadores_nuevos:

                # CÁLCULO MATEMÁTICO (La Resta):
                datos_nuevos = contadores_nuevos[NOMBRE_INTERFAZ]
                datos_antiguos = contadores_antiguos[NOMBRE_INTERFAZ]

                # Lo que hay ahora MENOS lo que había hace 1 segundo = Lo que pasó por el cable
                b_recv = datos_nuevos.bytes_recv - datos_antiguos.bytes_recv
                b_sent = datos_nuevos.bytes_sent - datos_antiguos.bytes_sent

                # Conversión a Megabits (Mbps) para saber la velocidad de internet
                mbps_bajada = (b_recv * 8) / 1_000_000
                mbps_subida = (b_sent * 8) / 1_000_000
                total_mbps = mbps_bajada + mbps_subida
                hora = datetime.datetime.now().strftime("%H:%M:%S") # Obtenemos la hora actual para el reporte

                # Colores Dinámicos:
                # Si la velocidad supera el límite, píntalo ROJO, si no, VERDE/AZUL.
                col_b = ROJO if mbps_bajada > UMBRAL_MBPS else VERDE
                col_s = ROJO if mbps_subida > UMBRAL_MBPS else AZUL
                
                # Imprimir en pantalla el estado actual
                print(f"[{hora}] Bajada: {col_b}{convertir_bytes(b_recv)}/s{RESET} | Subida: {col_s}{convertir_bytes(b_sent)}/s{RESET}")

                # --- SISTEMA DE ALERTA ---
                if total_mbps > UMBRAL_MBPS:
                    ahora = time.time()
                    if (ahora - ultima_alerta_bw) > TIEMPO_ENTRE_ALERTAS_BW:
                        print(f" {ROJO}[!] ALERTA ENVIADA AL CORREO{RESET}")
                        logging.warning(f"ALERTA TRÁFICO: {total_mbps:.2f} Mbps")
                        
                        # Usamos etiquetas HTML simples mezcladas con el texto
                        cuerpo = (f"<strong>ALERTA DE ANCHO DE BANDA CRÍTICO</strong>\n\n"
                                  f"Reportado por el equipo: {mi_ip}\n"
                                  f"Hora del evento: {hora}\n"
                                  f"-----------------------------------\n"
                                  f"Velocidad Total: <span style='color:red; font-weight:bold'>{total_mbps:.2f} Mbps</span>\n"
                                  f"Descarga: {mbps_bajada:.2f} Mbps\n"
                                  f"Subida: {mbps_subida:.2f} Mbps\n"
                                  f"-----------------------------------\n"
                                  f"Trabajo Final Fundamentos de Programacion\n"
                                  f"Grupo 3\n"
                                  f"Salon: 803\n"
                                  f"Carrera: Ingenieria de Redes y Comunicaciones\n"
                                  f"Universidad Peruana de Ciencias Aplicadas\n"
                                  f"<strong>Exigete, Innova, UPC</strong>")
                        
                        # Llamamos a la función que pone el logo al final
                        enviar_email_con_adjuntos(
    f"[ALERTA TRÁFICO CRÍTICO] {total_mbps:.0f} Mbps Detectados",
    cuerpo,
    []
)
                        
                        ultima_alerta_bw = ahora    # Reseteamos el reloj

                # Guardamos los datos en el cuaderno (CSV) para el gráfico final
                with open(ARCHIVO_CSV, mode='a', newline='') as file:
                    writer = csv.writer(file)
                    writer.writerow([hora, b_recv, b_sent])
                contadores_antiguos = contadores_nuevos # Actualizamos la referencia para el siguiente segundo
            else:
                print(f"Error: Interfaz {NOMBRE_INTERFAZ} no encontrada."); break
            
    # Si presionas Ctrl + C, salimos del bucle y preguntamos si quieres que se exporte el Grafico.                      
    except KeyboardInterrupt:
        print("\n[*] Deteniendo monitor...")
        if input("¿Generar gráfico? (s/n): ").lower() == 's': generar_grafico_trafico()
        return

def modulo():
    print("[*] Ejecutando módulo...")

    # lógica del módulo

    input("Presiona Enter para volver al menú...")
    return
# ==============================================================================
# 6. MÓDULO PING Y DISPONIBILIDAD - MONITOR INTELIGENTE
# ==============================================================================

# --- FUNCIONES AUXILIARES DE PING ---
def ping_cargar_servidores():
    servidores = []
    try:
        with open(PING_ARCHIVO_SERVIDORES, "r") as archivo:
            lector = csv.reader(archivo)
            for fila in lector:
                if len(fila) >= 2:
                    nombre = fila[0].strip(); ip = fila[1].strip()
                    if nombre and ip: servidores.append((nombre, ip))
    except IOError:
        servidores = PING_SERVIDORES_DEFECTO[:]
    return servidores

def ping_hacer_ping(ip_destino, timeout_ms=1000):
    sistema = platform.system().lower()
    if sistema == "windows":
        # -n 1 = un intento, -w = tiempo de espera en ms
        comando = f"ping -n 1 -w {timeout_ms} {ip_destino} > nul"
    else:
        # -c 1 = un intento, -W = tiempo de espera en segundos
        timeout_seg = max(1, int(timeout_ms / 1000))
        comando = f"ping -c 1 -W {timeout_seg} {ip_destino} > /dev/null 2>&1"
    
    # Ejecutamos el ping silencioso (sin output a consola)
    t1 = time.time()
    resultado = os.system(comando)
    t2 = time.time()
    
    if resultado == 0: return True, (t2 - t1) * 1000.0
    else: return False, None

def ping_guardar_historial(mediciones):
    existe = os.path.exists(PING_ARCHIVO_HISTORIAL)
    with open(PING_ARCHIVO_HISTORIAL, "a", newline="") as archivo:
        escritor = csv.writer(archivo)
        if not existe:
            escritor.writerow(["nombre", "ip", "intento", "timestamp", "exito", "rtt_ms"])
        for m in mediciones: escritor.writerow(m)

def ping_resumen_por_servidor(mediciones):
    resumen = {}
    for nombre, ip, _, _, exito, rtt_ms in mediciones:
        clave = (nombre, ip)
        if clave not in resumen:
            resumen[clave] = {"total":0, "exitos":0, "fallos":0, "suma_rtt":0.0}
        resumen[clave]["total"] += 1
        if exito:
            resumen[clave]["exitos"] += 1
            resumen[clave]["suma_rtt"] += rtt_ms
        else:
            resumen[clave]["fallos"] += 1
    return resumen

def ping_guardar_resumen_csv(resumen):
    with open(PING_ARCHIVO_RESUMEN, "w", newline="", encoding="utf-8") as archivo:
        writer = csv.writer(archivo)
        writer.writerow(["Servidor", "IP", "Total_Pings", "Exitos", "Fallos", "Disponibilidad_%", "RTT_Promedio_ms"])
        for (nombre, ip), datos in resumen.items():
            total = datos["total"]
            exitos = datos["exitos"]
            fallos = datos["fallos"]
            disp = (exitos * 100 / total) if total > 0 else 0
            rtt_prom = (datos["suma_rtt"] / exitos) if exitos > 0 else 0
            writer.writerow([nombre, ip, total, exitos, fallos, round(disp, 2), round(rtt_prom, 2)])

def ping_clasificar_estado(disponibilidad):
    if disponibilidad >= PING_UMBRAL_EXCELENTE: return f"{VERDE}EXCELENTE{RESET}"
    elif disponibilidad >= PING_UMBRAL_BUENO: return f"{VERDE}BUENA{RESET}"
    elif disponibilidad >= PING_UMBRAL_ACEPTABLE: return f"{AMARILLO}ACEPTABLE{RESET}"
    else: return f"{ROJO}CRÍTICA{RESET}"

def enviar_alerta_ping(nombre_servidor, ip_servidor):
    mi_ip = obtener_ip_local()
    cuerpo = (f"<strong>ALERTA DE DISPONIBILIDAD DE SERVICIO</strong>\n\n"
              f"Reportado por el monitor: {mi_ip}\n"
              f"-----------------------------------\n"
              f"Se ha detectado una caída de servicio crítica.\n\n"
              f"Servidor: <strong>{nombre_servidor}</strong>\n"
              f"IP Objetivo: <span style='color:red; font-weight:bold'>{ip_servidor}</span>\n"
              f"Estado: <span style='background-color:red; color:white; padding:2px'>OFFLINE / NO RESPONDE</span>\n"
              f"-----------------------------------\n"
              f"Acción sugerida: Verificar conectividad física o reiniciar el servicio.\n"
              f"Trabajo Final Fundamentos de Programacion\n"
              f"Grupo 3 - UPC")
    
    # Mensaje en consola para confirmar envío
    print(f" >> [ENVIANDO] Alerta crítica por correo a los administradores...")
    enviar_email_con_adjuntos(f"[ALERTA CRÍTICA] Servidor Caído - {nombre_servidor}", cuerpo, [])

# --- FUNCIONES PRINCIPALES DE PING ---

def ejecutar_ping_automatico():
    """
    Opción 3: HERRAMIENTA MANUAL.
    Permite al usuario ingresar una IP y cantidad de repeticiones para diagnóstico.
    """
    limpiar_pantalla()
    print(f"{NEGRITA}--- HERRAMIENTA DE DIAGNÓSTICO (PING MANUAL) ---{RESET}\n")

    # 1. Solicitamos la IP o Dominio
    target = input(f"Ingrese la IP o Dominio a probar (Ej: 8.8.8.8): {AMARILLO}")
    print(f"{RESET}", end="") # Resetear color
    
    if not target:
        print(f"{ROJO}Error: No ingresaste ninguna IP.{RESET}")
        input("Presiona Enter para volver..."); return

    # 2. Solicitamos la cantidad de repeticiones
    try:
        veces_input = input(f"¿Cuántas repeticiones? [Defecto: 4]: {AMARILLO}")
        print(f"{RESET}", end="")
        veces = int(veces_input) if veces_input else 4
    except ValueError:
        veces = 4 # Si escribe letras, usamos 4 por defecto

    print(f"\n{AZUL}Haciendo ping a {NEGRITA}{target}{RESET}{AZUL} con {veces} paquetes de datos:{RESET}\n")

    # Variables para estadísticas
    enviados = 0
    recibidos = 0
    tiempos = []

    try:
        for i in range(veces):
            enviados += 1
            # Usamos tu función auxiliar existente
            exito, rtt = ping_hacer_ping(target, timeout_ms=1000)

            if exito:
                recibidos += 1
                tiempos.append(rtt)
                print(f"Respuesta desde {target}: tiempo={VERDE}{int(rtt)}ms{RESET}")
            else:
                print(f"{ROJO}Tiempo de espera agotado para esta solicitud.{RESET}")

            time.sleep(1) # Pausa de 1 segundo entre pings (como Windows)

        # 3. Mostrar Estadísticas Finales
        perdidos = enviados - recibidos
        porcentaje_perdida = (perdidos / enviados) * 100

        print(f"\n{NEGRITA}--- Estadísticas de ping para {target} ---{RESET}")
        print(f"    Paquetes: enviados = {enviados}, recibidos = {recibidos}, perdidos = {perdidos} ({int(porcentaje_perdida)}% perdidos)")

        if tiempos:
            minimo = min(tiempos)
            maximo = max(tiempos)
            promedio = sum(tiempos) / len(tiempos)
            print(f"Tiempos aproximados de ida y vuelta en milisegundos:")
            print(f"    Mínimo = {int(minimo)}ms, Máximo = {int(maximo)}ms, Media = {AZUL}{int(promedio)}ms{RESET}")

    except KeyboardInterrupt:
        print(f"\n\n{AMARILLO}[!] Ping cancelado por usuario.{RESET}")
    
    input("\nPresiona Enter para volver al menú...")


def ejecutar_monitoreo_disponibilidad():
    """
    Opción 4: MONITOR INTELIGENTE (BUCLE INFINITO)
    - Silencioso (No imprime pings individuales).
    - Solo avisa si cambia el estado (Online <-> Offline).
    - Envía correo si cae un servidor.
    - Ctrl + C para salir.
    """
    limpiar_pantalla()
    print(f"{NEGRITA}--- MONITOR DE DISPONIBILIDAD (VIGILANCIA 24/7) ---{RESET}")
    print(f"{AMARILLO}[*] Presiona Ctrl + C para detener y volver al menú.{RESET}")
    print(f"[*] Iniciando vigilancia de servidores...\n")
    
    servidores = ping_cargar_servidores()
    if not servidores:
        print(f"{ROJO}No hay servidores configurados.{RESET}")
        input("Enter para volver..."); return

    # Diccionario para recordar el estado anterior de cada IP
    # Formato: { '192.168.1.50': 'ONLINE', ... }
    estado_previo = {}

    try:
        while True:
            # Iteramos por cada servidor
            for nombre, ip in servidores:
                exito, rtt = ping_hacer_ping(ip, timeout_ms=1000)
                
                # Determinamos el estado actual
                if exito:
                    estado_actual = "ONLINE"
                else:
                    estado_actual = "OFFLINE"

                # LÓGICA DE DETECCIÓN DE CAMBIOS
                if ip not in estado_previo:
                    # Primera vez que vemos este servidor: Mostramos estado inicial
                    color = VERDE if estado_actual == "ONLINE" else ROJO
                    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] Estado Inicial: {nombre} ({ip}) -> {color}{estado_actual}{RESET}")
                    estado_previo[ip] = estado_actual
                
                elif estado_previo[ip] != estado_actual:
                    # ¡CAMBIO DETECTADO! (El estado guardado es diferente al actual)
                    hora = datetime.datetime.now().strftime('%H:%M:%S')
                    
                    if estado_actual == "OFFLINE":
                        # CASO CRÍTICO: Se cayó el servidor
                        print(f"{ROJO}[{hora}] ¡ALERTA! {nombre} ({ip}) ha cambiado a OFFLINE{RESET}")
                        logging.warning(f"MONITOR: {nombre} ({ip}) CAÍDO (OFFLINE)")
                        enviar_alerta_ping(nombre, ip) # Enviar correo
                    else:
                        # CASO RECUPERACIÓN: Volvió a vivir
                        print(f"{VERDE}[{hora}] RESTAURADO: {nombre} ({ip}) ha vuelto a ONLINE{RESET}")
                        logging.warning(f"MONITOR: {nombre} ({ip}) RECUPERADO (ONLINE)")
                    
                    # Actualizamos el estado previo con el nuevo
                    estado_previo[ip] = estado_actual
            
            # Pausa para no saturar la red (Wait time)
            time.sleep(2)

    except KeyboardInterrupt:
        # Capturamos Ctrl + C para salir elegantemente sin romper el programa principal
        print(f"\n\n{AMARILLO}[!] Deteniendo monitor... Volviendo al menú principal.{RESET}")
        time.sleep(1)
        return

# ==============================================================================
# 7. MÓDULO ESCÁNER DE RED (Integración: Madai Gonzales)
# ==============================================================================
def enviar_alerta_desconocido(lista_desconocidos):
    """
    Envía un correo con la lista de intrusos detectados.
    """
    mi_ip = obtener_ip_local()
    
    # Construimos la lista HTML de dispositivos
    items_html = ""
    for ip, mac in lista_desconocidos:
        items_html += f"<li>IP: <strong>{ip}</strong> - MAC: {mac}</li>"

    cuerpo = (f"<strong>ALERTA DE SEGURIDAD - DISPOSITIVOS NO AUTORIZADOS</strong>\n\n"
              f"Reportado por: {mi_ip}\n"
              f"-----------------------------------\n"
              f"El escáner de red ha detectado dispositivos que NO están en la Lista Blanca:\n"
              f"<ul><span style='color:red; font-weight:bold'>{items_html}</span></ul>\n"
              f"-----------------------------------\n"
              f"Acción sugerida: Verificar si son invitados o intrusos.\n"
              f"-----------------------------------\n"
              f"Trabajo Final Fundamentos de Programacion\n"
              f"Grupo 3\n"
              f"Salon: 803\n"
              f"Carrera: Ingenieria de Redes y Comunicaciones\n"
              f"Universidad Peruana de Ciencias Aplicadas\n"
              f"<strong>Exigete, Innova, UPC</strong>")
    
    print(f" >> [ENVIANDO] Alerta de dispositivos desconocidos")
    
    # AQUÍ ESTÁ EL LOGO: Se pasa NOMBRE_LOGO como 3er argumento
    enviar_email_con_adjuntos(
    "[ALERTA] Intrusos en la Red Detectados",
    cuerpo,
    []
)

    print(f" >> [EXITO] Informe enviado al administrador.")

def ejecutar_scanner():
    limpiar_pantalla()
    print(f"{NEGRITA}--- MÓDULO: Escáner de Dispositivos (Lista Blanca) ---{RESET}")
    print(f"[*] Escaneando rango: {AMARILLO}{RANGO_RED_SCANNER}{RESET}")
    print("[*] Por favor espere, esto puede tardar unos segundos...\n")

    # Broadcast ARP
    arp = ARP(pdst=RANGO_RED_SCANNER)
    ether = Ether(dst="ff:ff:ff:ff:ff:ff")
    packet = ether/arp

    try:
        result = srp(packet, timeout=2, verbose=0, iface=NOMBRE_INTERFAZ)[0]
    except Exception as e:
        print(f"{ROJO}[ERROR] Falló el escaneo: {e}{RESET}")
        print("Verifique que el nombre de la interfaz sea correcto.")
        input("Enter para volver..."); return

    unknown_devices = []
    print(f"{'IP':<16} {'MAC':<20} {'ESTADO'}")
    print("-" * 60)

    for sent, received in result:
        ip = received.psrc
        mac = received.hwsrc
        
        # Lógica de Lista Blanca
        if ip in IP_MAC_CONFIANZA and IP_MAC_CONFIANZA[ip] == mac:
            estado = f"{VERDE}AUTORIZADO{RESET}"
        elif ip in IP_MAC_CONFIANZA and IP_MAC_CONFIANZA[ip] != mac:
            estado = f"{ROJO}CONFLICTO MAC{RESET}"
            unknown_devices.append((ip, mac))
        else:
            estado = f"{AMARILLO}DESCONOCIDO{RESET}"
            unknown_devices.append((ip, mac))

        print(f"{ip:<16} {mac:<20} {estado}")

    print("-" * 60)

    if unknown_devices:
        print(f"\n{ROJO}[!] Se detectaron {len(unknown_devices)} dispositivos sospechosos.{RESET}")
        for ip, mac in unknown_devices:
            logging.warning(f"INTRUSO DETECTADO: IP {ip} MAC {mac} no está en lista blanca.")
        
        print("Generando reporte por correo...")
        enviar_alerta_desconocido(unknown_devices) # Llamada a la función con correo+logo
    else:
        print(f"\n{VERDE}[OK] Todos los dispositivos encontrados son confiables.{RESET}")

    input("\nPresiona Enter para volver al menú...")
    return

# ==============================================================================
# 8. FUNCIÓN PRINCIPAL
# ==============================================================================

def generar_reporte_consolidado():
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill

    # 👇 ESTA LÍNEA ES OBLIGATORIA Y PRIMERA
    eventos = []

    # =============================
    # LOG PRINCIPAL
    # =============================
    if os.path.exists("seguridad_sistema.log"):
        with open("seguridad_sistema.log", "r", encoding="utf-8", errors="ignore") as f:
            for linea in f:
                try:
                    fecha_hora, detalle = linea.strip().split(" - ", 1)
                    fecha, hora = fecha_hora.split(" ")
                    eventos.append([fecha, hora, "SEGURIDAD", detalle])
                except:
                    pass

    # =============================
    # HISTORIAL PING
    # =============================
    if os.path.exists(PING_ARCHIVO_HISTORIAL):
        df_ping = pd.read_csv(PING_ARCHIVO_HISTORIAL)
        for _, r in df_ping.iterrows():
            fecha, hora = r['timestamp'].split(" ")
            tipo = "PING_FALLO" if not r['exito'] else "PING_OK"
            eventos.append([fecha, hora, tipo, f"{r['nombre']} ({r['ip']})"])

    # =============================
    # RESUMEN PING (CSV)
    # =============================
    if os.path.exists(PING_ARCHIVO_RESUMEN):
        df_ping_res = pd.read_csv(PING_ARCHIVO_RESUMEN)
        hoy = datetime.date.today().isoformat()

        for _, r in df_ping_res.iterrows():
            eventos.append([
                hoy,
                "00:00:00",
                "PING_RESUMEN",
                f"{r['Servidor']} ({r['IP']}) DISP:{r['Disponibilidad_%']}%"
            ])

    # =============================
    # HISTORIAL ANCHO DE BANDA
    # =============================
    if os.path.exists(ARCHIVO_CSV):
        df_bw = pd.read_csv(ARCHIVO_CSV)
        hoy = datetime.date.today().isoformat()
        for _, r in df_bw.iterrows():
            eventos.append([
                hoy,
                r['Tiempo'],
                "TRAFICO",
                f"Bajada:{r['Bytes_Bajada']} Subida:{r['Bytes_Subida']}"
            ])

    if not eventos:
        print("No existen eventos reales para consolidar.")
        return

    # =============================
    # DATAFRAME Y EXCEL
    # =============================
    df = pd.DataFrame(eventos, columns=["Fecha", "Hora", "Tipo_Alerta", "Detalle"])

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    excel_file = f"reporte_consolidado_{timestamp}.xlsx"

    with pd.ExcelWriter(excel_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Eventos", index=False)

        resumen = df['Tipo_Alerta'].value_counts().reset_index()
        resumen.columns = ["Tipo_Alerta", "Cantidad"]
        resumen.to_excel(writer, sheet_name="Resumen", index=False)

        fechas = df['Fecha'].value_counts().sort_index().reset_index()
        fechas.columns = ["Fecha", "Cantidad"]
        fechas.to_excel(writer, sheet_name="Eventos_por_Fecha", index=False)

    wb = load_workbook(excel_file)
    ws_res = wb["Resumen"]
    ws_fec = wb["Eventos_por_Fecha"]
    ws_dash = wb.create_sheet("Dashboard")
    # =============================
    # GRÁFICO PING (DASHBOARD)
    # =============================
    if os.path.exists(PING_ARCHIVO_RESUMEN):
        df_ping_res = pd.read_csv(PING_ARCHIVO_RESUMEN)

        # Crear hoja oculta de soporte (NO visible al usuario)
        ws_tmp = wb.create_sheet("_tmp_ping")

        ws_tmp.append([
            "Servidor",
            "RTT_Promedio_ms",
            "Pings_Recibidos",
            "Pings_Perdidos"
        ])

        for _, r in df_ping_res.iterrows():
            ws_tmp.append([
                r["Servidor"],
                r["RTT_Promedio_ms"],
                r["Exitos"],
                r["Fallos"]
            ])

        # --- Gráfico de barras (Recibidos / Perdidos) ---
        bar = BarChart()
        bar.title = "Métricas de Ping por Servidor"
        bar.y_axis.title = "Cantidad de Pings"
        bar.x_axis.title = "Servidor"

        data = Reference(
            ws_tmp,
            min_col=3,
            min_row=1,
            max_col=4,
            max_row=ws_tmp.max_row
        )
        cats = Reference(
            ws_tmp,
            min_col=1,
            min_row=2,
            max_row=ws_tmp.max_row
        )

        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)

        # --- Gráfico de línea (RTT promedio) ---
        line = LineChart()
        line.y_axis.title = "RTT Promedio (ms)"

        data_line = Reference(
            ws_tmp,
            min_col=2,
            min_row=1,
            max_row=ws_tmp.max_row
        )
        line.add_data(data_line, titles_from_data=True)
        line.set_categories(cats)

        # Combinar gráficos
        bar += line

        # Insertar en Dashboard (posición inferior)
        ws_dash.add_chart(bar, "B36")

        # Ocultar hoja temporal
        ws_tmp.sheet_state = "hidden"


    fondo = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    for row in ws_dash.iter_rows(min_row=1, max_row=40, min_col=1, max_col=20):
        for cell in row:
            cell.fill = fondo

    ws_dash.merge_cells("A1:J2")
    ws_dash["A1"] = "REPORTE DE INDICADORES DE SEGURIDAD – GRUPO 3"
    ws_dash["A1"].font = Font(size=20, bold=True)
    ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")

    bar = BarChart()
    bar.title = "Alertas por Tipo"
    data = Reference(ws_res, min_col=2, min_row=1, max_row=ws_res.max_row)
    cats = Reference(ws_res, min_col=1, min_row=2, max_row=ws_res.max_row)
    bar.add_data(data, titles_from_data=True)
    bar.set_categories(cats)
    ws_dash.add_chart(bar, "B4")

    line = LineChart()
    line.title = "Eventos en el Tiempo"
    data = Reference(ws_fec, min_col=2, min_row=1, max_row=ws_fec.max_row)
    cats = Reference(ws_fec, min_col=1, min_row=2, max_row=ws_fec.max_row)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    ws_dash.add_chart(line, "B20")

    wb.save(excel_file)

    os.startfile(excel_file)

    enviar_email_con_adjuntos(
        "REPORTE DE INDICADORES DE SEGURIDAD – GRUPO 3",
        "Reporte generado únicamente con eventos reales del sistema.",
        [excel_file]
    )

def main():
    try:
        crear_caratula(NOMBRE_LOGO)
        input(f"{AMARILLO}Presione ENTER para continuar al menú principal...{RESET}")
        limpiar_pantalla()

        while True:
            limpiar_pantalla()
            mostrar_menu()
            opcion = input("Opción: ").strip()

            if opcion == "1":
                ejecutar_arp_sniffer()

            elif opcion == "2":
                monitor_ancho_banda()

            elif opcion == "3":
                ejecutar_ping_automatico()

            elif opcion == "4":
                ejecutar_monitoreo_disponibilidad()

            elif opcion == "5":
                ejecutar_scanner()

            elif opcion == "6":
                generar_reporte_consolidado()

            elif opcion == "7":
                print("Saliendo del sistema...")
                break

            else:
                print("Opción inválida")
                time.sleep(1)

    except KeyboardInterrupt:
        print(f"\n{AMARILLO}Ejecución interrumpida por el usuario. Saliendo del sistema...{RESET}")

if __name__ == "__main__":
    main()
